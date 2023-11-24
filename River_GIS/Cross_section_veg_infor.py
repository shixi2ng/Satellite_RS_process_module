# coding=utf-8
from osgeo import gdal
import os
import numpy as np
import pandas as pd
import geopandas as gd
import basic_function as bf
import sys
from Aborted_codes import Landsat_main_v1 as ls
from matplotlib import pyplot as plt
from pylab import mpl
import openpyxl
from openpyxl.chart import LineChart, Reference, ScatterChart, Series


### Description:
### This code is used to generate the 1-Dimensional distribution of required information alongside the cross-section.
### Mainly contains two classes: (1) Mosaic raster of different tiles.
###                              (2) Equalise divided cross-sections.
### The specific function under each class will be described in each class.
### All Rights Reserved

class Mosaic_Raster(object):
    # A bundle of Mosaic Raster documented in Dic
    def __init__(self, rasterfile_path, mosaic_raster_list):
        """

        :param rasterfile_path:
        :param mosaic_raster_list:
        """
        # Initialisation
        self.rasterfile_path = rasterfile_path
        self.rasterfile = bf.file_filter(self.rasterfile_path, ['.tif', '.TIF'], and_or_factor='or',
                                         subfolder_detection=True,
                                         exclude_word_list=['.xml', '.aux', '.cpg', '.dbf', '.lock'])
        self.mosaic_raster_list = mosaic_raster_list
        # Type check for mosaic raster list
        if type(self.mosaic_raster_list) is not list:
            print('Please input the mosaic raster list as a list')
            sys.exit(-1)
        else:
            try:
                temp = self.mosaic_raster_list[0][0]
            except:
                print('Please check whether the combined raster list is under the list with two dimension')
                sys.exit(-1)

        try:
            self.combined_dic = {'raster_name': [temp[0] for temp in self.mosaic_raster_list]}
            for raster_name_temp in self.combined_dic['raster_name']:
                self.combined_dic[raster_name_temp] = [temp[1:] for temp in self.mosaic_raster_list if
                                                       raster_name_temp in temp][0]
        except:
            print('Mosaic raster initialisation error!')
            sys.exit(-1)

        if self.combined_dic['raster_name'] == []:
            print('combined raster list is empty!')
            sys.exit(-1)

    def mosaic_raster(self, output_path, overwritten_factor=False):
        """

        :param output_path:
        """
        bf.create_folder(output_path)
        bf.check_file_path(output_path)
        for raster_name_temp in self.combined_dic['raster_name']:
            if not os.path.exists(output_path + raster_name_temp + '.tif') or (overwritten_factor and os.path.exists(output_path + raster_name_temp + '.tif')):
                list_temp = []
                for filename in self.combined_dic[raster_name_temp]:
                    list_len_ori = len(list_temp)
                    list_temp.extend([temp for temp in self.rasterfile if filename in temp])
                    if len(list_temp) - list_len_ori > 1:
                        print('There is more than one file named ' + str(
                            filename) + '. Please change the file name in the combined raster list!')
                        sys.exit(-1)
                    elif len(list_temp) - list_len_ori == 0:
                        print('There is no file with name ' + str(
                            filename) + '. Please double check the file name in the combined raster list!')
                        sys.exit(-1)
                ds_temp = gdal.Open(list_temp[0])
                ulx, xres_temp, xskew, uly, yskew, yres_temp = ds_temp.GetGeoTransform()
                nodatavalue = ds_temp.GetRasterBand(1).GetNoDataValue()
                gdal_temp = gdal.Warp(output_path + raster_name_temp + '.tif', list_temp, format="GTiff", options=["COMPRESS=LZW", "TILED=YES"], xRes=xres_temp, yRes=yres_temp, srcNodata=np.nan, dstNodata=np.nan, outputType=gdal.GDT_Float32)
                gdal_temp = None


class Equalised_Polyline_Raster(object):
    def __init__(self, shpfile_path, raster_path, unit=1, information_tag='Veg', date_tag=None, preprocess_factor=True):
        """

        :type shpfile_path: object
        """
        self.equalised_dic = {}
        self.cross_section_list = []
        self.shpfile = shpfile_path
        self.raster_path = raster_path
        self.unit = unit
        self.information_tag = information_tag
        self.date_tag = date_tag
        self.integrate_dem_factor = False
        self.Fig_output = {}
        self.xlsx_output_filepath = None
        self.missing_DEM_list = []
        self.redundant_DEM_list = []

        if self.date_tag is None:
            self.date_tag = ''
            print('Caution! The date tag is not input!')

        if not os.path.exists(self.shpfile):
            print('Shpfile does not exist! Double check!')
            sys.exit(-1)

        if not os.path.exists(self.raster_path):
            print('Raster file does not exist! Double check!')
            sys.exit(-1)

        self.raster_src = gdal.Open(raster_path)
        self.raster_array = bf.file2raster(raster_path)
        self.ulx, xres, xskew, self.uly, yskew, yres = self.raster_src.GetGeoTransform()
        self.lrx = self.ulx + (self.raster_src.RasterXSize * xres)
        self.lry = self.uly + (self.raster_src.RasterYSize * yres)
        self.raster_temp = self.raster_src.GetRasterBand(1).ReadAsArray()

        # Pre-Process raster file
        if preprocess_factor:
            self.raster_array = self.raster_array.astype(np.float)
            self.raster_array[self.raster_array == -32768] = np.nan
            self.raster_array = self.raster_array / 10000


    def polyline2equalpoint(self, cross_section_col='cross_sect'):
        """

        :param unit:
        :param cross_section_col:
        """
        cross_gd_temp = gd.read_file(self.shpfile)
        cross_geometry = cross_gd_temp[[cross_section_col, 'geometry']]
        self.cross_section_list = cross_gd_temp[cross_section_col]
        for i in range(cross_geometry.shape[0]):
            cross_section_name = cross_geometry.loc[i][cross_section_col]
            cross_section_geometry = cross_geometry.loc[i]['geometry'].coords.xy
            left_bank_coord = [cross_section_geometry[0][0], cross_section_geometry[1][0]]
            right_bank_coord = [cross_section_geometry[0][-1], cross_section_geometry[1][-1]]
            distance = np.sqrt(
                (left_bank_coord[0] - right_bank_coord[0]) ** 2 + (left_bank_coord[1] - right_bank_coord[1]) ** 2)
            size = int(np.ceil(distance / self.unit))
            self.equalised_dic[cross_section_name] = np.array([[itr * self.unit, left_bank_coord[0] + (right_bank_coord[0] - left_bank_coord[0]) * itr / size, left_bank_coord[1] + (right_bank_coord[1] - left_bank_coord[1]) * itr / size] for itr in range(size + 1)])

    def retrieve_inf4point(self):
        """

        :return:
        """
        if not bool(self.equalised_dic):
            print('Please equalise the polyline')
            return

        for section_name in self.cross_section_list:
            section_coord_temp = self.equalised_dic[section_name]
            section_information_temp = np.array(
                [bf.query_with_cor(self.raster_src, section_coord_temp[i, 1], section_coord_temp[i, 2], srcnanvalue=0, dstnanvalue=np.nan, raster=self.raster_array) for i in
                 range(section_coord_temp.shape[0])]).transpose()
            section_information_temp = section_information_temp.reshape([section_information_temp.shape[0], 1])
            if section_coord_temp.shape[0] == section_information_temp.shape[0]:
                self.equalised_dic[section_name + '_' + self.information_tag] = np.concatenate([section_coord_temp, section_information_temp], axis=1)
                self.equalised_dic[section_name + '_' + self.information_tag] = np.delete(self.equalised_dic[section_name + '_' + self.information_tag], np.argwhere(np.isnan(self.equalised_dic[section_name + '_' + self.information_tag])), axis=0)
                self.equalised_dic[section_name + '_' + self.information_tag] = pd.DataFrame(self.equalised_dic[section_name + '_' + self.information_tag],
                                                       columns=[self.information_tag + '点距左岸距离/' + str(self.unit) + 'm', '经度',
                                                                '纬度', self.information_tag])
            else:
                print('calcuation error!')
                sys.exit(-1)

    def xlsx_output(self, output_path, file_name, overwritten_factor=False, plot_figure_in_excel=True):
        """

        :param plot_figure_in_excel:
        :param overwritten_factor:
        :param output_path:
        :param file_name:
        """
        bf.check_file_path(output_path)

        if '.' in file_name:
            if 'xlsx' not in file_name:
                print('Please input the file name with xlsx extension')
        else:
            file_name = file_name + '.xlsx'

        self.xlsx_output_filepath = output_path + file_name
        if not bool(list(self.cross_section_list)) or not bool(self.equalised_dic):
            print('Please implement the polyline2equalpoint process before!')
            return

        if not os.path.exists(output_path + file_name) or (overwritten_factor and os.path.exists(output_path + file_name)):
            writer = pd.ExcelWriter(output_path + file_name, engine='xlsxwriter')
            for section_name in self.cross_section_list:
                cross_section_output_df = self.equalised_dic[section_name + '_' + self.information_tag]
                cross_section_output_df.to_excel(writer, sheet_name=str(section_name))
            writer.save()

        if plot_figure_in_excel:
            wb_temp = openpyxl.load_workbook(self.xlsx_output_filepath)
            for sheet_name_temp in wb_temp.sheetnames:
                chart_temp = None
                chart_temp2 = None
                sheet_temp = None

                sheet_temp = wb_temp[sheet_name_temp]

                if sheet_temp.max_column == 5:
                    chart_temp = ScatterChart(scatterStyle='lineMarker')
                    # chart_temp.layout = Layout(manualLayout=ManualLayout(x=0.1, y=0.1, h=0.8, w=0.8))
                    chart_temp.height = 15
                    chart_temp.width = sheet_temp.max_row / 40
                    x_value_1 = Reference(sheet_temp, min_col=2, min_row=2,max_col=2, max_row=sheet_temp.max_row)
                    y_value_1 = Reference(sheet_temp, min_col=5, min_row=2,max_col=5, max_row=sheet_temp.max_row)
                    series_1 = Series(y_value_1, x_value_1, title=self.information_tag)
                    chart_temp.series.append(series_1)
                    chart_temp.y_axis.title = self.information_tag

                    chart_temp.y_axis.tickLblPos = "low"
                    chart_temp.x_axis.title = '距左岸距离/m'
                    chart_temp.x_axis.tickLblPos = "low"
                    chart_temp.x_axis.axPos = 'b'
                    chart_temp.y_axis.axPos = 'r'
                    chart_temp.title = self.information_tag + '沿横断面变化'

                    sheet_temp.add_chart(chart_temp, 'G2')
                elif sheet_temp.max_column == 7:
                    chart_temp = ScatterChart(scatterStyle='lineMarker')
                    # chart_temp.layout = Layout(manualLayout=ManualLayout(x=0.1, y=0.1, h=0.8, w=0.8))
                    chart_temp.height = 15
                    chart_temp.width = sheet_temp.max_row / 40
                    x_value_1 = Reference(sheet_temp, min_col=2, min_row=2, max_col=2, max_row=sheet_temp.max_row)
                    y_value_1 = Reference(sheet_temp, min_col=5, min_row=2, max_col=5, max_row=sheet_temp.max_row)
                    series_1 = Series(y_value_1, x_value_1, title=self.information_tag)
                    chart_temp.series.append(series_1)
                    chart_temp.y_axis.title = self.information_tag
                    chart_temp.y_axis.majorGridlines = None
                    # chart_temp.y_axis.tickLblPos = "low"
                    chart_temp.x_axis.title = '距左岸距离/m'
                    chart_temp.x_axis.tickLblPos = "low"


                    chart_temp2 = ScatterChart(scatterStyle='lineMarker')
                    x_value_2 = Reference(sheet_temp, min_col=6, min_row=2, max_col=6, max_row=sheet_temp.max_row)
                    y_value_2 = Reference(sheet_temp, min_col=7, min_row=2, max_col=7, max_row=sheet_temp.max_row)
                    series_2 = Series(y_value_2, x_value_2, title='高程')
                    chart_temp2.series.append(series_2)
                    chart_temp2.y_axis.tickLblPos = "low"
                    chart_temp2.y_axis.title = '高程/m'
                    chart_temp2.y_axis.axPos = 'l'
                    chart_temp2.y_axis.axId = 200

                    chart_temp.x_axis.crosses = 'min'
                    chart_temp.y_axis.crosses = 'max'
                    chart_temp.x_axis.axPos = 't'
                    chart_temp.y_axis.axPos = 'r'
                    chart_temp += chart_temp2
                    chart_temp.title = self.information_tag + '与高程沿横断面变化'

                    # s1 = chart_temp.series[0]
                    # s1.marker.symbol = "circle"
                    s2 = chart_temp2.series[0]
                    s2.marker.symbol = "circle"

                    sheet_temp.add_chart(chart_temp, 'I2')

            wb_temp.save(self.xlsx_output_filepath)

    def plot_cross_section(self, output_path):
        plt.rc('font', size=16)
        plt.rc('axes', linewidth=1.5)
        bf.check_file_path(output_path)
        for section_name in self.cross_section_list:
            try:
                temp = self.equalised_dic[section_name + '_' + self.information_tag]
            except:
                print('Please equalise before plot figure!')
                return
            dis_temp = temp[:, 0]
            data_temp = temp[:, 3]
            itr = 0
            while itr < data_temp.shape[0]:
                if np.isnan(data_temp[itr]):
                    dis_temp = np.delete(dis_temp, itr, 0)
                    data_temp = np.delete(data_temp, itr, 0)
                    itr -= 1
                itr += 1
            self.Fig_output['Fig_' + section_name], self.Fig_output['axe_' + section_name] = plt.subplots(figsize=(temp.shape[0]*12/1000, 6), constrained_layout=True)
            self.Fig_output['axe_' + section_name].plot(dis_temp, data_temp, color=(0/256,0/256,0/256), linewidth=2)
            self.Fig_output['axe_' + section_name].set_xlabel('距离/' + str(self.unit) + '(m)', fontsize=24, fontweight='bold')
            self.Fig_output['axe_' + section_name].set_ylabel(self.information_tag, fontsize=24, fontweight='bold')
            self.Fig_output['Fig_' + section_name].savefig(output_path + self.date_tag + '_' + str(section_name) + '_' + self.information_tag + '.png', dpi=300)
            plt.close('all')

    def thesholdlised(self, value, novel_inform_tag=None):
        if novel_inform_tag is None:
            novel_inform_tag = self.information_tag

    def integrate_DEM(self, dem_file_path):
        ### Please notice that the DEM file must with the specific header and format
        ### For instance:
        ### (1) DEM for all the sections should be placed in a same sheet
        ### (2) For each cross-section the information should start with the header like:
        ###     line1    |Name      |  Null                    |  Null    |
        ###     line2    |Year      |  Null                    |  Null    |
        ###     line3    |Series    |  Null                    |  Null    |
        ###     line4    |Series No |  Distance from left bank |  Height  |

        if not bool(self.equalised_dic):
            print('Please run the polyline2equalpoint process before!')
            return

        if not os.path.exists(dem_file_path):
            print('Please input a correct dem file path!')
            return

        try:
            DEM_table = pd.read_excel(dem_file_path)
            DEM_array = np.array(DEM_table)
        except:
            print('Something went wrong during the input of DEM')
            return

        for cs in self.cross_section_list:
            cs_initial_pos = np.argwhere(DEM_array[:, 0].reshape([DEM_array.shape[0], 1]) == cs)
            if cs_initial_pos.shape[0] == 0:
                self.missing_DEM_list.append(cs)
            elif cs_initial_pos.shape[0] > 1:
                self.redundant_DEM_list.append(cs)
            else:
                if int(DEM_array[cs_initial_pos[0,0] + 4, cs_initial_pos[0,1]]) == 1:
                    temp_factor = True
                    itr = 0
                    while temp_factor:
                        try:
                            temp_v = int(DEM_array[cs_initial_pos[0,0] + 4 + itr, cs_initial_pos[0,1]])
                            itr += 1
                        except:
                            temp_factor = False

                    try:
                        cs_dis_temp = DEM_array[cs_initial_pos[0, 0] + 4 : cs_initial_pos[0, 0] + 4 + itr, 1].astype(np.float)
                        cs_dem_temp = DEM_array[cs_initial_pos[0, 0] + 4 : cs_initial_pos[0, 0] + 4 + itr, 2].astype(np.float)
                    except:
                        print('Something went wrong during indexing the distance and dem data for ' + str(cs) + '!')

                    if cs + '_' + self.information_tag in self.equalised_dic.keys() and self.equalised_dic[cs + '_' + self.information_tag].size != 0:
                        self.equalised_dic[cs + '_' + self.information_tag]['高程点距左岸距离'] = pd.DataFrame(cs_dis_temp)
                        self.equalised_dic[cs + '_' + self.information_tag]['高程'] = pd.DataFrame(cs_dem_temp)
                        self.integrate_dem_factor = True
                    else:
                        print('Please retrieve inf4point firstly for ' + str(cs))

                else:
                    print('Please make sure the header for each cross section is coherence and correct in ' + str(cs))

    def cla(self):
        self.raster_array = None
        self.raster_temp = None
        self.raster_src = None
        self.equalised_dic = None
        self.Fig_output = None


if __name__ == "__main__":
    mpl.rcParams['font.sans-serif'] = ['SimSun']  # 指定默认字体
    mpl.rcParams['axes.unicode_minus'] = False
    gdal.UseExceptions()
    np.seterr(divide='ignore', invalid='ignore')
    root_path = "E:\\A_PhD_Main_stuff\\2022_04_22_Mid_Yangtze\\Sample_Landsat\\"
    original_file_path = root_path + 'Original_zipfile\\'
    corrupted_file_path = root_path + 'Corrupted\\'
    unzipped_file_path = root_path + 'Landsat_Ori_TIFF\\'
    mndwi_output_path = root_path + 'MNDWI_Mosaic\\'
    ndvi_output_path = root_path + 'NDVI_Mosaic\\'
    shpfile_path = 'E:\\A_PhD_Main_stuff\\2022_04_22_Mid_Yangtze\\Shpfile\\'
    shpfile_path_list = [shpfile_path + 'Zone_37_WGS84.shp', shpfile_path + 'Zone_38_WGS84.shp', shpfile_path + 'Zone_39_WGS84.shp']
    DEM_filepath = 'E:\\A_PhD_Main_stuff\\2022_04_22_Mid_Yangtze\\DEM\\2019_MD_YTZ_DEM.xlsx'
    S2_filepath = 'E:\\A_PhD_Main_stuff\\2022_04_22_Mid_Yangtze\\Sample_Sentinel\\Sentinel2_L2A_output\\Sentinel2_L2A_output\\'


    ls.create_folder(unzipped_file_path)
    file_metadata = ls.generate_landsat_metadata(original_file_path, unzipped_file_path,
                                                 corrupted_file_path, root_path, unzipped_para=False)
    ls.generate_landsat_vi(root_path, unzipped_file_path, file_metadata, vi_construction_para=True,
                                        construction_overwritten_para=False, cloud_removal_para=True,
                                        vi_clipped_para=False,
                                        clipped_overwritten_para=False, construct_dc_para=False,
                                        dc_overwritten_para=False,
                                        construct_sdc_para=False, sdc_overwritten_para=False,
                                        VI_list=['NDVI', 'MNDWI'])
    # mosaic_raster_list = [['2020_prior_fl_S2', '20200427_49RFN', '20200427_49RGN', '20200429_50RKT', '20200429_50RKU', '20200429_50RKV', '20200429_50RLT', '20200429_50RLU', '20200429_50RMT', '20200429_50RMU', '20200502_49REP', '20200502_49RFP', '20200502_49RGP']]
    # mr_NDVI = Mosaic_Raster(S2_filepath + 'NDVI\\', mosaic_raster_list)
    # mr_MNDWI = Mosaic_Raster(S2_filepath + 'NDWI\\', mosaic_raster_list)
    # mr_NDVI.mosaic_raster(ndvi_output_path)
    # mr_MNDWI.mosaic_raster(mndwi_output_path)


    epr_dic = {}
    vi_output_folder = {}
    for shp in shpfile_path_list:
        for VI in ['MNDWI']:
            vi_output_folder[VI] = root_path + VI + '_equal_inform\\'
            bf.create_folder(vi_output_folder[VI])
            for raster_temp in bf.file_filter(root_path + VI + '_Mosaic\\', ['.tif','2020'], and_or_factor='and', exclude_word_list=['aux']):
                ppf = True
                if 'S2' in raster_temp:
                    ppf = False
                shpfile_name_temp = shp[shp.find('Zone_3'): shp.find('Zone_3') + 7]
                date_tag_temp = raster_temp[raster_temp.find('_Mosaic') + 8: raster_temp.find('.tif')]
                epr_dic[shpfile_name_temp + '_' + date_tag_temp] = Equalised_Polyline_Raster(shp, raster_temp, information_tag=VI, date_tag=date_tag_temp, preprocess_factor=ppf)
                epr_dic[shpfile_name_temp + '_' + date_tag_temp].polyline2equalpoint()
                epr_dic[shpfile_name_temp + '_' + date_tag_temp].retrieve_inf4point()
                epr_dic[shpfile_name_temp + '_' + date_tag_temp].integrate_DEM(DEM_filepath)
                bf.create_folder(vi_output_folder[VI] + 'Table2\\')
                bf.create_folder(vi_output_folder[VI] + 'Figure\\')
                epr_dic[shpfile_name_temp + '_' + date_tag_temp].xlsx_output(vi_output_folder[VI] + 'Table2\\', date_tag_temp + '_' + shpfile_name_temp + '_' + VI + '.xlsx', plot_figure_in_excel=True)
                # epr_dic[shpfile_name_temp + '_' + date_tag_temp].plot_cross_section(vi_output_folder[VI] + 'Figure\\')
                epr_dic[shpfile_name_temp + '_' + date_tag_temp].cla()